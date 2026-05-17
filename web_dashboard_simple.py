#!/usr/bin/env python3
"""
Job Hunter Dashboard - Complete Simple Version
NO LOGIN REQUIRED - Direct job display
"""

import os
from datetime import date
from flask import Flask, render_template_string, request

app = Flask(__name__)

# Complete HTML template with NO login
HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Job Hunter Dashboard</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body { 
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; 
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
            min-height: 100vh; 
            padding: 20px;
        }
        .container { max-width: 1400px; margin: 0 auto; }
        
        /* Header */
        .header { 
            background: white; 
            border-radius: 20px; 
            padding: 30px; 
            margin-bottom: 30px; 
            text-align: center;
            box-shadow: 0 10px 30px rgba(0,0,0,0.1);
        }
        .header h1 { color: #333; margin-bottom: 10px; font-size: 32px; }
        .header p { color: #666; }
        
        /* Stats Cards */
        .stats { 
            display: grid; 
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); 
            gap: 20px; 
            margin-bottom: 30px; 
        }
        .stat-card { 
            background: white; 
            border-radius: 15px; 
            padding: 25px; 
            text-align: center; 
            box-shadow: 0 5px 15px rgba(0,0,0,0.1);
            transition: transform 0.2s;
        }
        .stat-card:hover { transform: translateY(-5px); }
        .stat-card .number { font-size: 36px; font-weight: bold; color: #667eea; }
        .stat-card .label { color: #666; margin-top: 10px; }
        
        /* Filters */
        .filters { 
            background: white; 
            border-radius: 15px; 
            padding: 15px 20px; 
            margin-bottom: 20px; 
            display: flex; 
            gap: 10px; 
            flex-wrap: wrap;
            align-items: center;
        }
        .filter-btn { 
            padding: 8px 20px; 
            border: 2px solid #e0e0e0; 
            background: white; 
            border-radius: 25px; 
            cursor: pointer; 
            font-size: 14px;
            transition: all 0.2s;
        }
        .filter-btn:hover { border-color: #667eea; }
        .filter-btn.active { background: #667eea; color: white; border-color: #667eea; }
        
        /* Search Bar */
        .search-bar { 
            background: white; 
            border-radius: 15px; 
            padding: 15px 20px; 
            margin-bottom: 20px; 
            display: flex; 
            gap: 10px; 
        }
        .search-bar input { 
            flex: 1; 
            padding: 12px; 
            border: 2px solid #e0e0e0; 
            border-radius: 10px; 
            font-size: 16px; 
        }
        .search-bar button { 
            padding: 12px 30px; 
            background: #667eea; 
            color: white; 
            border: none; 
            border-radius: 10px; 
            cursor: pointer; 
            font-size: 16px;
        }
        .search-bar button:hover { background: #5a67d8; }
        
        /* Job Table */
        .job-table { 
            background: white; 
            border-radius: 20px; 
            overflow-x: auto; 
            box-shadow: 0 5px 15px rgba(0,0,0,0.1); 
        }
        table { width: 100%; border-collapse: collapse; min-width: 600px; }
        th { 
            background: #667eea; 
            color: white; 
            padding: 15px; 
            text-align: left; 
            font-weight: 600; 
        }
        td { padding: 15px; border-bottom: 1px solid #f0f0f0; }
        tr:hover { background: #f8f9ff; }
        
        .score-high { color: #10b981; font-weight: bold; }
        .score-mid { color: #f59e0b; font-weight: bold; }
        .score-low { color: #ef4444; font-weight: bold; }
        
        .status { 
            display: inline-block; 
            padding: 4px 12px; 
            border-radius: 20px; 
            font-size: 12px; 
            font-weight: 600;
        }
        .status-new { background: #dbeafe; color: #2563eb; }
        .status-applied { background: #d1fae5; color: #059669; }
        .status-interview { background: #fef3c7; color: #d97706; }
        
        .btn { 
            padding: 6px 14px; 
            border: none; 
            border-radius: 8px; 
            cursor: pointer; 
            font-size: 12px; 
            text-decoration: none; 
            display: inline-block;
            margin: 2px;
        }
        .btn-primary { background: #667eea; color: white; }
        .btn-success { background: #10b981; color: white; }
        .btn-primary:hover, .btn-success:hover { opacity: 0.8; }
        
        .job-link { color: #667eea; text-decoration: none; }
        .job-link:hover { text-decoration: underline; }
        
        .empty-state { 
            text-align: center; 
            padding: 60px; 
            background: white; 
            border-radius: 20px; 
        }
        .empty-state h3 { color: #333; margin-bottom: 10px; }
        .empty-state p { color: #666; }
        
        @media (max-width: 768px) {
            body { padding: 10px; }
            th, td { padding: 10px; font-size: 12px; }
            .btn { padding: 4px 8px; font-size: 10px; }
            .stat-card .number { font-size: 24px; }
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>🎯 Job Hunter Dashboard</h1>
            <p>AI-powered job matching • Tailored CVs • Smart tracking</p>
        </div>
        
        <div class="stats">
            <div class="stat-card">
                <div class="number">{{ stats.total }}</div>
                <div class="label">Total Jobs</div>
            </div>
            <div class="stat-card">
                <div class="number">{{ stats.high_score }}</div>
                <div class="label">80%+ Matches</div>
            </div>
            <div class="stat-card">
                <div class="number">{{ stats.applied }}</div>
                <div class="label">Applied</div>
            </div>
            <div class="stat-card">
                <div class="number">{{ stats.today }}</div>
                <div class="label">Found Today</div>
            </div>
        </div>
        
        <div class="filters">
            <button class="filter-btn active" data-filter="all">All Jobs</button>
            <button class="filter-btn" data-filter="high">80%+ Matches</button>
            <button class="filter-btn" data-filter="applied">Applied</button>
            <button class="filter-btn" data-filter="new">New</button>
        </div>
        
        <div class="search-bar">
            <input type="text" id="searchInput" placeholder="Search by job title or company..." onkeyup="searchJobs()">
            <button onclick="searchJobs()">🔍 Search</button>
        </div>
        
        <div class="job-table">
            <table id="jobTable">
                <thead>
                    <tr>
                        <th>#</th>
                        <th>Score</th>
                        <th>Job Title</th>
                        <th>Company</th>
                        <th>Platform</th>
                        <th>Status</th>
                        <th>Actions</th>
                    </tr>
                </thead>
                <tbody id="jobTableBody">
                    {% for job in jobs %}
                    <tr data-score="{{ job.score }}" data-status="{{ job.status }}" data-title="{{ job.title|lower }}" data-company="{{ job.company|lower }}">
                        <td>{{ loop.index }}</td>
                        <td class="score-{% if job.score >= 80 %}high{% elif job.score >= 60 %}mid{% else %}low{% endif %}">{{ job.score }}%</td>
                        <td><a href="{{ job.link }}" target="_blank" class="job-link">{{ job.title[:80] }}</a></td>
                        <td>{{ job.company[:50] }}</td>
                        <td>{{ job.platform }}</td>
                        <td><span class="status status-{{ job.status }}">{{ job.status }}</span></td>
                        <td>
                            <a href="{{ job.link }}" target="_blank" class="btn btn-primary">View</a>
                            <button class="btn btn-success" onclick="generateCV('{{ job.link }}', '{{ job.title|escapejs }}')">CV</button>
                        </td>
                    </tr>
                    {% endfor %}
                </tbody>
            </table>
        </div>
        
        {% if not jobs %}
        <div class="empty-state">
            <h3>📭 No Jobs Found</h3>
            <p>Run your job scraper to find matching positions.</p>
            <p style="margin-top: 15px;">💡 Run <code>python3 scraper.py</code> in your terminal</p>
        </div>
        {% endif %}
    </div>
    
    <script>
        // Filter jobs
        document.querySelectorAll('.filter-btn').forEach(btn => {
            btn.addEventListener('click', function() {
                const filter = this.dataset.filter;
                const rows = document.querySelectorAll('#jobTableBody tr');
                
                rows.forEach(row => {
                    const score = parseInt(row.dataset.score);
                    const status = row.dataset.status;
                    
                    if (filter === 'all') row.style.display = '';
                    else if (filter === 'high') row.style.display = score >= 80 ? '' : 'none';
                    else if (filter === 'applied') row.style.display = status === 'Applied' ? '' : 'none';
                    else if (filter === 'new') row.style.display = status === 'New' ? '' : 'none';
                });
                
                document.querySelectorAll('.filter-btn').forEach(b => b.classList.remove('active'));
                this.classList.add('active');
            });
        });
        
        // Search jobs
        function searchJobs() {
            const searchTerm = document.getElementById('searchInput').value.toLowerCase();
            const rows = document.querySelectorAll('#jobTableBody tr');
            
            rows.forEach(row => {
                const title = row.dataset.title;
                const company = row.dataset.company;
                if (title.includes(searchTerm) || company.includes(searchTerm)) {
                    row.style.display = '';
                } else {
                    row.style.display = 'none';
                }
            });
        }
        
        // Generate CV
        function generateCV(link, title) {
            alert('📄 CV Generation\n\nJob: ' + title + '\n\nSend this link to your Telegram bot to generate a tailored CV:\n' + link);
        }
    </script>
</body>
</html>
"""

def get_jobs():
    """Load jobs from your Excel file"""
    jobs = []
    today = date.today().strftime("%Y-%m-%d")
    
    try:
        from openpyxl import load_workbook
        tracker_file = os.path.expanduser("~/Desktop/JobHunter/job_tracker.xlsx")
        
        if os.path.exists(tracker_file):
            wb = load_workbook(tracker_file, data_only=True)
            
            if "Master Tracker" in wb.sheetnames:
                ws = wb["Master Tracker"]
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row and len(row) >= 4 and row[0] and row[1] and row[2] and row[3]:
                        # Parse score
                        score_str = str(row[1]).replace('%', '')
                        try:
                            score = int(float(score_str))
                        except:
                            score = 0
                        
                        # Only show jobs with score >= 60
                        if score >= 60:
                            jobs.append({
                                'title': str(row[3])[:200] if row[3] else 'Unknown',
                                'company': str(row[2])[:100] if row[2] else 'Unknown',
                                'platform': str(row[4])[:50] if row[4] else 'Unknown',
                                'link': str(row[7]) if row[7] and len(str(row[7])) > 10 else '#',
                                'score': score,
                                'status': 'New'
                            })
        
        # Sort by score (highest first)
        jobs.sort(key=lambda x: -x['score'])
        
    except Exception as e:
        print(f"Error loading jobs: {e}")
    
    return jobs

@app.route('/')
def dashboard():
    jobs = get_jobs()
    
    stats = {
        'total': len(jobs),
        'high_score': len([j for j in jobs if j['score'] >= 80]),
        'applied': 0,
        'today': 0
    }
    
    return render_template_string(HTML_TEMPLATE, jobs=jobs, stats=stats)

@app.route('/health')
def health():
    return 'OK'

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    print(f"\n🚀 Job Hunter Dashboard Running on port {port}")
    print(f"📍 Open: http://localhost:{port}\n")
    app.run(host='0.0.0.0', port=port)
