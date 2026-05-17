#!/usr/bin/env python3
"""
Telegram Listener — Mohammed Job Hunter
Runs in background, listens for cv commands, generates and sends CV + Cover Letter
Usage: python3 telegram_listener.py
"""

import os
import sys
import time
import json
import requests
import subprocess
from datetime import date

TELEGRAM_TOKEN = "8972917892:AAGs_Z6xWc67poi7EfVdJpPoJJb_3hs8sJo"
TELEGRAM_CHAT_ID = "8872960522"
CV_BUILDER = os.path.expanduser("~/Desktop/JobHunter/cv_builder.py")
CVS_DIR = os.path.expanduser("~/Desktop/JobHunter/CVs")
TRACKER_FILE = os.path.expanduser("~/Desktop/JobHunter/job_tracker.xlsx")
SCRAPER_PATH = os.path.expanduser("~/Desktop/jobhunter/scraper.py")
CREDS_FILE = os.path.expanduser("~/Desktop/jobhunter/credentials.json")
SHEET_ID = os.environ.get("SHEET_ID", "")

# Store today's job list in memory
todays_jobs = []
last_update_id = 0

def send_message(text):
    try:
        url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
        data = {
            "chat_id": TELEGRAM_CHAT_ID,
            "text": text,
            "parse_mode": "HTML",
            "disable_web_page_preview": True
        }
        requests.post(url, json=data, timeout=10)
    except Exception as e:
        print(f"Send error: {e}")

def send_file(filepath, caption=""):
    try:
        url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendDocument"
        with open(filepath, 'rb') as f:
            files = {'document': f}
            data = {
                'chat_id': TELEGRAM_CHAT_ID,
                'caption': caption,
                'parse_mode': 'HTML'
            }
            requests.post(url, files=files, data=data, timeout=30)
        print(f"Sent file: {filepath}")
    except Exception as e:
        print(f"File send error: {e}")

def get_updates(offset=0):
    try:
        url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/getUpdates"
        params = {"offset": offset, "timeout": 30, "allowed_updates": ["message"]}
        response = requests.get(url, params=params, timeout=35)
        if response.status_code == 200:
            return response.json().get("result", [])
    except Exception as e:
        print(f"Get updates error: {e}")
    return []

def load_todays_jobs():
    """Load today's jobs from master tracker"""
    global todays_jobs
    today = date.today().strftime("%Y-%m-%d")
    jobs = []

    try:
        from openpyxl import load_workbook
        if not os.path.exists(TRACKER_FILE):
            return

        wb = load_workbook(TRACKER_FILE, data_only=True)
        if "Master Tracker" not in wb.sheetnames:
            return

        ws = wb["Master Tracker"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and str(row[0]).strip() == today:
                jobs.append({
                    "num": len(jobs) + 1,
                    "date": str(row[0]),
                    "score": str(row[1]),
                    "company": str(row[2]) if row[2] else "",
                    "title": str(row[3]) if row[3] else "",
                    "platform": str(row[4]) if row[4] else "",
                    "salary": str(row[5]) if row[5] else "TBD",
                    "seen_before": str(row[6]) if row[6] else "",
                    "link": str(row[7]) if row[7] else "",
                })

        todays_jobs = jobs
        print(f"✅ Loaded {len(jobs)} jobs from today's tracker")

    except Exception as e:
        print(f"Load jobs error: {e}")

def generate_cv_for_job(job_num):
    """Generate CV and cover letter for job number"""
    global todays_jobs

    # Reload jobs from tracker in case it was updated
    load_todays_jobs()

    if not todays_jobs:
        send_message("❌ No jobs found for today. Run the scraper first!")
        return

    # Find job by number
    job = None
    for j in todays_jobs:
        if j['num'] == job_num:
            job = j
            break

    if not job:
        send_message(f"❌ Job #{job_num} not found. Today has {len(todays_jobs)} jobs (1-{len(todays_jobs)})")
        return

    link = job.get('link', '')
    if not link or link == 'None':
        send_message(f"❌ No link found for job #{job_num}")
        return

    send_message(
        f"⏳ Generating CV for job #{job_num}...\n\n"
        f"📋 <b>{job['title']}</b>\n"
        f"🏢 {job['company']}\n\n"
        f"This takes about 30-60 seconds..."
    )

    try:
        # Run cv_builder
        result = subprocess.run(
            [sys.executable, CV_BUILDER, link],
            capture_output=True, text=True, timeout=120,
            cwd=os.path.expanduser("~/Desktop/JobHunter")
        )

        if result.returncode != 0:
            send_message(f"❌ CV builder error:\n{result.stderr[-500:]}")
            return

        # Find generated files
        company_clean = ''.join(c for c in job['company'] if c.isalnum() or c == ' ').strip().replace(' ', '_')
        title_clean = ''.join(c for c in job['title'] if c.isalnum() or c == ' ').strip().replace(' ', '_')

        # Look for recently created files in CVs folder
        if not os.path.exists(CVS_DIR):
            send_message("❌ CVs folder not found")
            return

        cv_files = []
        for f in os.listdir(CVS_DIR):
            if f.endswith('.docx') and not f.startswith('_'):
                full_path = os.path.join(CVS_DIR, f)
                # Get files modified in last 3 minutes
                if time.time() - os.path.getmtime(full_path) < 180:
                    cv_files.append(full_path)

        if not cv_files:
            send_message("❌ Could not find generated files. Check your Mac.")
            return

        # Send files
        send_message(f"✅ CV generated for <b>{job['title']}</b> at <b>{job['company']}</b>!\n\nSending files...")

        cv_sent = False
        cl_sent = False

        for filepath in cv_files:
            filename = os.path.basename(filepath)
            if filename.startswith('CV_'):
                send_file(filepath, f"📄 CV — {job['title']} at {job['company']}")
                cv_sent = True
            elif filename.startswith('CoverLetter_'):
                send_file(filepath, f"📝 Cover Letter — {job['title']} at {job['company']}")
                cl_sent = True

        if cv_sent or cl_sent:
            send_message(
                f"🎉 Done! Files sent above.\n\n"
                f"📋 Job: {job['title']}\n"
                f"🏢 Company: {job['company']}\n"
                f"🔗 Apply: {link}\n\n"
                f"Don't forget to mark ✅ in your Excel tracker after applying!"
            )

            # Mark CV as generated in tracker
            mark_cv_generated(job_num)
        else:
            send_message("⚠️ Files generated but could not send. Check your CVs folder on Mac.")

    except subprocess.TimeoutExpired:
        send_message("⏰ CV generation timed out. Try again or check your Mac.")
    except Exception as e:
        send_message(f"❌ Error: {str(e)}")

def mark_cv_generated(job_num):
    """Mark CV as generated in master tracker"""
    try:
        from openpyxl import load_workbook
        today = date.today().strftime("%Y-%m-%d")

        wb = load_workbook(TRACKER_FILE)
        ws = wb["Master Tracker"]

        today_count = 0
        for row in ws.iter_rows(min_row=2):
            if row[0].value and str(row[0].value).strip() == today:
                today_count += 1
                if today_count == job_num:
                    row[9].value = f"✅ {date.today().strftime('%d/%m')}"
                    break

        wb.save(TRACKER_FILE)
    except Exception as e:
        print(f"Mark CV error: {e}")

def run_scraper():
    """Run scraper.py and notify via Telegram"""
    send_message("⏳ Running scraper... This takes 2-4 minutes.")
    try:
        result = subprocess.run(
            [sys.executable, SCRAPER_PATH],
            capture_output=True, text=True, timeout=300,
            cwd=os.path.expanduser("~/Desktop/jobhunter")
        )
        if result.returncode == 0:
            send_message("✅ Scraper done! Refresh your dashboard to see new jobs.")
        else:
            send_message(f"❌ Scraper error:\n{result.stderr[-400:]}")
    except subprocess.TimeoutExpired:
        send_message("⏰ Scraper timed out after 5 minutes.")
    except Exception as e:
        send_message(f"❌ Could not start scraper: {e}")

def check_dashboard_trigger():
    """Check Google Sheets Config tab for a RUN trigger from the dashboard"""
    if not SHEET_ID:
        return
    try:
        import gspread
        from oauth2client.service_account import ServiceAccountCredentials
        scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
        creds = ServiceAccountCredentials.from_json_keyfile_name(CREDS_FILE, scope)
        client = gspread.authorize(creds)
        spreadsheet = client.open_by_key(SHEET_ID)
        config = spreadsheet.worksheet("Config")
        trigger = config.cell(1, 2).value
        if trigger == "RUN":
            print("🚀 Dashboard trigger detected! Running scraper...")
            config.update('B1', [['RUNNING']])
            run_scraper()
            config.update('B1', [['IDLE']])
    except Exception as e:
        print(f"Trigger check error: {e}")

def handle_message(text):
    """Process incoming Telegram messages"""
    text = text.strip().lower()

    # CV command: "cv 3" or "cv3" or just "3"
    if text.startswith('cv'):
        # Extract number
        num_part = text.replace('cv', '').strip()
        try:
            job_num = int(num_part)
            generate_cv_for_job(job_num)
        except ValueError:
            send_message("❌ Please specify a job number.\nExample: <code>cv 3</code>")

    elif text.isdigit():
        job_num = int(text)
        generate_cv_for_job(job_num)

    elif text in ['run', '/run', 'scrape', '/scrape', 'scraper']:
        send_message("🔍 Starting scraper from Telegram command...")
        run_scraper()

    elif text in ['help', '/help', '/start']:
        send_message(
            "🤖 <b>Mohammed Job Hunter Bot</b>\n\n"
            "Commands:\n"
            "• <code>run</code> — Start the job scraper now\n"
            "• <code>cv 3</code> — Generate CV for job #3\n"
            "• <code>jobs</code> — Show today's job list\n"
            "• <code>help</code> — Show this message\n\n"
            "You can also click <b>Run Scraper</b> on the dashboard!"
        )

    elif text in ['jobs', '/jobs']:
        load_todays_jobs()
        if not todays_jobs:
            send_message("❌ No jobs found for today. Run the scraper!")
        else:
            msg = f"📋 <b>Today's Jobs ({len(todays_jobs)}):</b>\n\n"
            for j in todays_jobs[:10]:
                msg += f"#{j['num']} — {j['score']} — {j['title']} @ {j['company']}\n"
            if len(todays_jobs) > 10:
                msg += f"\n...and {len(todays_jobs)-10} more"
            send_message(msg)

    else:
        send_message(
            f"❓ Unknown command: <code>{text}</code>\n\n"
            "Try:\n"
            "• <code>cv 3</code> to generate CV for job #3\n"
            "• <code>jobs</code> to see today's list\n"
            "• <code>help</code> for all commands"
        )

def main():
    global last_update_id

    print("🤖 Mohammed Job Hunter Bot — Listener Started")
    print(f"📂 CVs folder: {CVS_DIR}")
    print(f"📊 Tracker: {TRACKER_FILE}")
    print("👂 Listening for commands... (Press Ctrl+C to stop)\n")

    send_message(
        "🤖 <b>Job Hunter Bot is listening!</b>\n\n"
        "Send <code>cv [number]</code> to generate CV\n"
        "Example: <code>cv 3</code>\n\n"
        "Send <code>jobs</code> to see today's list"
    )

    # Load today's jobs on startup
    load_todays_jobs()

    trigger_check_counter = 0

    while True:
        try:
            updates = get_updates(offset=last_update_id + 1)

            for update in updates:
                last_update_id = update["update_id"]

                if "message" in update and "text" in update["message"]:
                    text = update["message"]["text"]
                    chat_id = str(update["message"]["chat"]["id"])

                    # Only respond to Mohammed's chat
                    if chat_id == TELEGRAM_CHAT_ID:
                        print(f"📨 Received: {text}")
                        handle_message(text)
                    else:
                        print(f"⚠️ Message from unknown chat: {chat_id}")

            # Check dashboard trigger every ~30 seconds (15 loops × 2s)
            trigger_check_counter += 1
            if trigger_check_counter >= 15:
                trigger_check_counter = 0
                check_dashboard_trigger()

            time.sleep(2)

        except KeyboardInterrupt:
            print("\n👋 Bot stopped.")
            send_message("🔴 Job Hunter Bot stopped.")
            break
        except Exception as e:
            print(f"Loop error: {e}")
            time.sleep(5)

if __name__ == "__main__":
    main()